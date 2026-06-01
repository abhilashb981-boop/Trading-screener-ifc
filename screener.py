"""
╔══════════════════════════════════════════════════════════════╗
║          PROFESSIONAL SELL SIGNAL SCREENER v2.0              ║
║    Multi-Indicator Mathematical Trading System               ║
╚══════════════════════════════════════════════════════════════╝

Indicators Used (11-Indicator Scoring System):
  1. RSI          — Overbought Detection (>70 zone)
  2. MACD         — Bearish Crossover Signal
  3. Bollinger    — Upper Band Rejection
  4. Supertrend   — Trend Flip to Bearish
  5. EMA Cross    — 9 EMA crosses below 21 EMA
  6. Volume       — Above-average on red candle
  7. Stochastic   — %K crosses below %D in overbought zone
  8. CISD         — Displacement candle + level break
  9. Candle       — Bearish engulfing / shooting star
 10. ADX          — Trend strength confirmation
 11. IFC          — Institutional Funding Candle (swing high rejection)

Usage:
  python trading_screener.py --input stocks.xlsx --exchange NSE
"""

import argparse
import warnings
warnings.filterwarnings('ignore')

import pandas as pd
import numpy as np
import yfinance as yf
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════
#                  INDICATOR CALCULATIONS
# ══════════════════════════════════════════════════════

def calc_rsi(close, period=14):
    delta = close.diff()
    gain = delta.clip(lower=0).rolling(period).mean()
    loss = (-delta.clip(upper=0)).rolling(period).mean()
    rs = gain / loss.replace(0, np.nan)
    return 100 - (100 / (1 + rs))


def calc_macd(close, fast=12, slow=26, signal=9):
    ema_fast = close.ewm(span=fast, adjust=False).mean()
    ema_slow = close.ewm(span=slow, adjust=False).mean()
    macd_line = ema_fast - ema_slow
    signal_line = macd_line.ewm(span=signal, adjust=False).mean()
    histogram = macd_line - signal_line
    return macd_line, signal_line, histogram


def calc_bollinger(close, period=20, std_dev=2):
    ma = close.rolling(period).mean()
    std = close.rolling(period).std()
    upper = ma + (std_dev * std)
    lower = ma - (std_dev * std)
    percent_b = (close - lower) / (upper - lower)
    return upper, ma, lower, percent_b


def calc_supertrend(df, period=10, multiplier=3.0):
    atr = calc_atr(df, period)
    hl2 = (df['High'] + df['Low']) / 2
    
    upper_basic = hl2 + (multiplier * atr)
    lower_basic = hl2 - (multiplier * atr)
    
    upper_band = upper_basic.copy()
    lower_band = lower_basic.copy()
    
    for i in range(1, len(df)):
        upper_band.iloc[i] = min(upper_basic.iloc[i], upper_band.iloc[i-1]) \
            if df['Close'].iloc[i-1] <= upper_band.iloc[i-1] else upper_basic.iloc[i]
        lower_band.iloc[i] = max(lower_basic.iloc[i], lower_band.iloc[i-1]) \
            if df['Close'].iloc[i-1] >= lower_band.iloc[i-1] else lower_basic.iloc[i]
    
    supertrend = pd.Series(index=df.index, dtype=float)
    direction = pd.Series(index=df.index, dtype=int)  # 1=bullish, -1=bearish
    
    for i in range(1, len(df)):
        if df['Close'].iloc[i] <= upper_band.iloc[i]:
            supertrend.iloc[i] = upper_band.iloc[i]
            direction.iloc[i] = -1
        else:
            supertrend.iloc[i] = lower_band.iloc[i]
            direction.iloc[i] = 1
    
    return supertrend, direction


def calc_atr(df, period=14):
    high = df['High']
    low = df['Low']
    close = df['Close']
    tr = pd.concat([
        high - low,
        (high - close.shift(1)).abs(),
        (low - close.shift(1)).abs()
    ], axis=1).max(axis=1)
    return tr.rolling(period).mean()


def calc_stochastic(df, k_period=14, d_period=3):
    lowest_low = df['Low'].rolling(k_period).min()
    highest_high = df['High'].rolling(k_period).max()
    stoch_k = 100 * (df['Close'] - lowest_low) / (highest_high - lowest_low)
    stoch_d = stoch_k.rolling(d_period).mean()
    return stoch_k, stoch_d


def calc_adx(df, period=14):
    high = df['High']
    low = df['Low']
    close = df['Close']
    
    plus_dm = high.diff()
    minus_dm = -low.diff()
    plus_dm[plus_dm < 0] = 0
    minus_dm[minus_dm < 0] = 0
    
    mask = plus_dm < minus_dm
    plus_dm[mask] = 0
    mask2 = minus_dm < plus_dm
    minus_dm[mask2] = 0
    
    atr = calc_atr(df, period)
    plus_di = 100 * (plus_dm.rolling(period).mean() / atr)
    minus_di = 100 * (minus_dm.rolling(period).mean() / atr)
    
    dx = 100 * (plus_di - minus_di).abs() / (plus_di + minus_di)
    adx = dx.rolling(period).mean()
    return adx, plus_di, minus_di


def detect_bearish_candle(df):
    """Bearish Engulfing or Shooting Star detect చేయడం"""
    curr = df.iloc[-1]
    prev = df.iloc[-2]
    
    curr_body = abs(curr['Close'] - curr['Open'])
    curr_range = curr['High'] - curr['Low']
    upper_wick = curr['High'] - max(curr['Open'], curr['Close'])
    
    # Shooting Star: long upper wick, small body at bottom
    shooting_star = (
        upper_wick >= 2 * curr_body and
        curr['Close'] < curr['Open'] and
        curr_body > 0
    )
    
    # Bearish Engulfing: current red candle engulfs previous green candle
    bearish_engulf = (
        prev['Close'] > prev['Open'] and        # prev green
        curr['Close'] < curr['Open'] and        # curr red
        curr['Open'] >= prev['Close'] and       # opens above prev close
        curr['Close'] <= prev['Open']           # closes below prev open
    )
    
    return shooting_star or bearish_engulf, "Shooting Star" if shooting_star else ("Bearish Engulfing" if bearish_engulf else "None")


def detect_cisd(df, atr, lookback=20, multiplier=2.0):
    """CISD Break detect చేయడం"""
    recent = df.tail(lookback)
    recent_atr = atr.tail(lookback)
    
    for i in range(2, len(recent) - 1):
        candle = recent.iloc[i]
        curr_atr = recent_atr.iloc[i]
        if pd.isna(curr_atr) or curr_atr == 0:
            continue
        
        body = candle['Close'] - candle['Open']
        is_spike = body > multiplier * curr_atr and candle['Close'] > candle['Open']
        
        if is_spike:
            cisd_level = candle['Open']
            for j in range(i + 1, len(recent)):
                if recent.iloc[j]['Close'] < cisd_level:
                    return True
    
    return False


def detect_institutional_funding_candle(df, atr, swing_lookback=40):
    """
    Institutional Funding Candle (IFC) Pattern detect చేయడం
    
    Pattern Logic (image నుంచి):
      Step 1 → Recent Swing High identify చేయడం (resistance level)
      Step 2 → Price ఆ level కింద consolidate అయిందా చూడటం (zigzag)
      Step 3 → Institutions "fund" చేశారా — green candles తో swing high వరకు push
      Step 4 → Swing high దగ్గర bearish reversal — red candle(s) తో reject అయిందా
    
    Returns: (bool, detail_string)
    """
    if len(df) < swing_lookback + 5:
        return False, "⚪ Insufficient bars"

    atr_val = atr.iloc[-1]
    if pd.isna(atr_val) or atr_val <= 0:
        return False, "⚪ No ATR"

    recent = df.tail(swing_lookback).reset_index(drop=True)
    n = len(recent)

    # ── Step 1: Swing High ──
    # Last 5 candles को exclude చేసి swing high find చేయడం
    # (recent rally ముందు formed అయిన high)
    swing_window = recent.iloc[:n - 5]
    if swing_window.empty:
        return False, "⚪ No swing window"

    swing_high_pos = swing_window['High'].idxmax()
    swing_high     = swing_window['High'].max()

    # Swing high meaningful గా ఉండాలి (current price కంటే ATR లో significant గా high)
    curr_close = recent.iloc[-1]['Close']
    if swing_high < curr_close - 0.5 * atr_val:
        return False, "⚪ Swing high below current price"

    # ── Step 2: Consolidation below swing high ──
    # Swing high తర్వాత price కింద ఉండిన candles >= 3
    after_swing = recent.iloc[swing_high_pos + 1 : n - 3]
    if len(after_swing) < 3:
        return False, "⚪ No consolidation window"

    below_count = (after_swing['Close'] < swing_high - 0.5 * atr_val).sum()
    if below_count < 3:
        return False, "⚪ No clear consolidation below swing high"


    # ── Step 3: Funding rally — swing high వైపు consecutive green candles ──
    # Last 10 candles లో swing high కి దగ్గరగా వెళ్ళే rally ఉండాలి.
    # Criteria:
    #   a) కనీసం 2 green candles in rally window
    #   b) rally window లో highest High >= swing_high - 2.0 * ATR
    rally_zone   = recent.iloc[max(0, n - 10) : n - 1]
    green_count  = int((rally_zone['Close'] > rally_zone['Open']).sum())
    funding_high = float(rally_zone['High'].max())

    rally_reached_swing = funding_high >= swing_high - 2.0 * atr_val

    if green_count < 2 or not rally_reached_swing:
        return False, (f"⚪ No funding rally (greens={green_count}, "
                       f"high={funding_high:.1f}, need>={swing_high - 2*atr_val:.1f})")
    funding_found = True   # used below


    # ── Step 4: Bearish rejection FROM swing high region ──
    # Two criteria, either triggers a valid IFC:
    #   A) Rally window లో high >= swing_high - 0.5 ATR touched
    #      AND last 2 candles లో red candle ఉంది (started reversing)
    #   B) Rally window లో swing high touched
    #      AND current close < rally high by > 1.0 ATR (already dropped)

    touch_high = funding_high   # from step 3

    # Condition A: touched + recent red candle
    recent_reds  = int((recent.iloc[n-3:n]['Close'] < recent.iloc[n-3:n]['Open']).sum())
    touched_swing = touch_high >= swing_high - 0.5 * atr_val

    # Condition B: price has already dropped from the touch
    curr_close    = recent.iloc[-1]['Close']
    already_dropped = (touch_high - curr_close) >= 1.0 * atr_val

    reversal_found = touched_swing and (recent_reds >= 1 or already_dropped)
    touch_high_val = touch_high

    if not reversal_found:
        if touch_high >= swing_high - 1.5 * atr_val:
            return False, f"⚪ Near swing high ₹{swing_high:.1f} — watch for reversal"
        return False, f"⚪ Rally didn't touch swing high (rally_high={touch_high:.1f}, swing={swing_high:.1f})"

    # ── All 4 conditions met ──
    detail = (f"Swing High: ₹{swing_high:.1f} | "
              f"Rally High: ₹{touch_high_val:.1f} | "
              f"Reversal: {recent_reds} red candle(s) after touch")
    return True, detail


# ══════════════════════════════════════════════════════
#                  MAIN ANALYSIS ENGINE
# ══════════════════════════════════════════════════════

def analyze_stock(symbol, period='3mo'):
    """
    Stock analyze చేసి sell score return చేయడం
    Returns: dict with all signals and score
    """
    try:
        ticker = yf.Ticker(symbol)
        df = ticker.history(period=period, interval='1d')
        
        if df.empty or len(df) < 30:
            return None, "Insufficient data"
        
        df = df[['Open', 'High', 'Low', 'Close', 'Volume']].copy()
        df.dropna(inplace=True)
        
        close = df['Close']
        score = 0
        signals = {}
        
        # ── 1. RSI ──
        rsi = calc_rsi(close)
        rsi_val = rsi.iloc[-1]
        rsi_prev = rsi.iloc[-2]
        rsi_signal = rsi_val > 65 and rsi_val < rsi_prev  # Overbought + falling
        if rsi_val > 70:
            score += 1.5
            signals['RSI'] = f"🔴 {rsi_val:.1f} (Overbought+Falling)"
        elif rsi_val > 65:
            score += 1.0
            signals['RSI'] = f"🟡 {rsi_val:.1f} (Elevated)"
        else:
            signals['RSI'] = f"⚪ {rsi_val:.1f} (Neutral)"
        
        # ── 2. MACD ──
        macd_line, signal_line, histogram = calc_macd(close)
        macd_cross = macd_line.iloc[-2] > signal_line.iloc[-2] and macd_line.iloc[-1] < signal_line.iloc[-1]
        macd_neg_hist = histogram.iloc[-1] < 0 and histogram.iloc[-1] < histogram.iloc[-2]
        if macd_cross:
            score += 1.5
            signals['MACD'] = "🔴 Bearish Crossover (Fresh)"
        elif macd_neg_hist:
            score += 0.75
            signals['MACD'] = "🟡 Histogram Declining"
        else:
            signals['MACD'] = "⚪ No Signal"
        
        # ── 3. Bollinger Bands ──
        upper, mid, lower, pct_b = calc_bollinger(close)
        pct_b_val = pct_b.iloc[-1]
        bb_rejection = close.iloc[-2] >= upper.iloc[-2] and close.iloc[-1] < upper.iloc[-1]
        if bb_rejection or pct_b_val > 0.95:
            score += 1.0
            signals['Bollinger'] = f"🔴 Upper Band Rejection ({pct_b_val:.2f})"
        elif pct_b_val > 0.80:
            score += 0.5
            signals['Bollinger'] = f"🟡 Near Upper Band ({pct_b_val:.2f})"
        else:
            signals['Bollinger'] = f"⚪ {pct_b_val:.2f}"
        
        # ── 4. Supertrend ──
        supertrend, direction = calc_supertrend(df)
        st_dir_curr = direction.iloc[-1]
        st_dir_prev = direction.iloc[-2]
        st_flip = st_dir_curr == -1 and st_dir_prev == 1
        if st_flip:
            score += 1.5
            signals['Supertrend'] = "🔴 Just Flipped BEARISH"
        elif st_dir_curr == -1:
            score += 0.5
            signals['Supertrend'] = "🟡 Bearish Trend"
        else:
            signals['Supertrend'] = "⚪ Bullish Trend"
        
        # ── 5. EMA Cross ──
        ema9 = close.ewm(span=9, adjust=False).mean()
        ema21 = close.ewm(span=21, adjust=False).mean()
        ema_cross = ema9.iloc[-2] > ema21.iloc[-2] and ema9.iloc[-1] < ema21.iloc[-1]
        ema_below = ema9.iloc[-1] < ema21.iloc[-1]
        if ema_cross:
            score += 1.5
            signals['EMA(9/21)'] = "🔴 Death Cross (Fresh)"
        elif ema_below:
            score += 0.5
            signals['EMA(9/21)'] = "🟡 9 EMA Below 21 EMA"
        else:
            signals['EMA(9/21)'] = "⚪ Bullish Alignment"
        
        # ── 6. Volume ──
        vol = df['Volume']
        vol_avg = vol.rolling(20).mean()
        is_red = close.iloc[-1] < df['Open'].iloc[-1]
        vol_surge = vol.iloc[-1] > vol_avg.iloc[-1] * 1.5
        if is_red and vol_surge:
            score += 1.0
            signals['Volume'] = f"🔴 High Vol Red Candle ({vol.iloc[-1]/vol_avg.iloc[-1]:.1f}x avg)"
        elif vol_surge:
            score += 0.0
            signals['Volume'] = f"🟡 Volume Surge ({vol.iloc[-1]/vol_avg.iloc[-1]:.1f}x avg)"
        else:
            signals['Volume'] = f"⚪ Normal ({vol.iloc[-1]/vol_avg.iloc[-1]:.1f}x avg)"
        
        # ── 7. Stochastic ──
        stoch_k, stoch_d = calc_stochastic(df)
        k_val = stoch_k.iloc[-1]
        d_val = stoch_d.iloc[-1]
        stoch_cross = stoch_k.iloc[-2] > stoch_d.iloc[-2] and stoch_k.iloc[-1] < stoch_d.iloc[-1]
        stoch_ob = k_val > 80 and stoch_cross
        if stoch_ob:
            score += 1.0
            signals['Stochastic'] = f"🔴 Overbought Crossover K:{k_val:.0f} D:{d_val:.0f}"
        elif k_val > 80:
            score += 0.5
            signals['Stochastic'] = f"🟡 Overbought K:{k_val:.0f}"
        else:
            signals['Stochastic'] = f"⚪ K:{k_val:.0f} D:{d_val:.0f}"
        
        # ── 8. CISD Break ──
        atr = calc_atr(df)
        cisd_found = detect_cisd(df, atr)
        if cisd_found:
            score += 1.5
            signals['CISD'] = "🔴 CISD Level Broken"
        else:
            signals['CISD'] = "⚪ No CISD Pattern"
        
        # ── 9. Candle Pattern ──
        bearish_candle, candle_name = detect_bearish_candle(df)
        if bearish_candle:
            score += 1.0
            signals['Candle'] = f"🔴 {candle_name}"
        else:
            curr = df.iloc[-1]
            body_pct = abs(curr['Close'] - curr['Open']) / (curr['High'] - curr['Low'] + 0.001) * 100
            signals['Candle'] = f"⚪ Regular ({body_pct:.0f}% body)"
        
        # ── 10. ADX (Trend Strength) ──
        adx, plus_di, minus_di = calc_adx(df)
        adx_val = adx.iloc[-1]
        trending_bear = adx_val > 25 and minus_di.iloc[-1] > plus_di.iloc[-1]
        if trending_bear:
            score += 1.0
            signals['ADX'] = f"🔴 Strong Bear Trend ({adx_val:.0f})"
        elif adx_val > 20:
            score += 0.25
            signals['ADX'] = f"🟡 Trending ({adx_val:.0f})"
        else:
            signals['ADX'] = f"⚪ Weak Trend ({adx_val:.0f})"
        
        # ── 11. Institutional Funding Candle (IFC) ──
        ifc_found, ifc_detail = detect_institutional_funding_candle(df, atr)
        if ifc_found:
            score += 2.0
            signals['IFC'] = f"🔴 IFC Pattern! {ifc_detail}"
        else:
            signals['IFC'] = ifc_detail if ifc_detail.startswith("⚪") else f"⚪ {ifc_detail}"

        # ── Score Normalize (0–10) ──
        max_possible = 14.0  # 12 original + 2 IFC
        final_score = round(min((score / max_possible) * 10, 10), 1)
        
        # ── Entry, SL, Target Calculation ──
        atr_val = atr.iloc[-1]
        entry = round(close.iloc[-1], 2)
        stop_loss = round(df['High'].tail(5).max() + 0.5 * atr_val, 2)
        target1 = round(entry - 1.5 * atr_val, 2)
        target2 = round(entry - 3.0 * atr_val, 2)
        risk = stop_loss - entry
        reward1 = entry - target1
        rr_ratio = round(reward1 / risk, 2) if risk > 0 else 0
        
        # ── Final Signal ──
        if final_score >= 7.0:
            signal = "🔴 STRONG SELL"
        elif final_score >= 5.0:
            signal = "🟠 SELL"
        elif final_score >= 3.5:
            signal = "🟡 WEAK SELL"
        else:
            signal = "⚪ NO SIGNAL"
        
        try:
            info = ticker.info
            company = info.get('longName', symbol)[:28]
            sector = info.get('sector', 'N/A')
        except:
            company = symbol
            sector = 'N/A'
        
        return {
            'symbol': symbol.replace('.NS', '').replace('.BO', ''),
            'ifc': ifc_found,
            'company': company,
            'sector': sector,
            'signal': signal,
            'score': final_score,
            'entry': entry,
            'stop_loss': stop_loss,
            'target1': target1,
            'target2': target2,
            'rr_ratio': rr_ratio,
            'atr': round(atr_val, 2),
            'rsi': round(rsi_val, 1),
            'signals': signals,
            'scan_date': datetime.now().strftime('%d-%m-%Y'),
        }, None
        
    except Exception as e:
        return None, str(e)


# ══════════════════════════════════════════════════════
#                  EXCEL OUTPUT
# ══════════════════════════════════════════════════════

def create_excel_report(sell_stocks, hold_stocks, output_path):
    wb = Workbook()
    
    # ── Colors & Styles ──
    C = {
        'bg_dark':    "0D0D0D",
        'bg_panel':   "141414",
        'red_strong': "FF2244",
        'red_soft':   "FF6B6B",
        'orange':     "FF8C00",
        'yellow':     "FFD700",
        'green':      "00C853",
        'grey':       "888888",
        'white':      "FFFFFF",
        'row_odd':    "141A14",
        'row_even':   "0F150F",
        'header_bg':  "1A0A0A",
    }
    
    def fill(color): return PatternFill("solid", start_color=color)
    def font(color, bold=False, size=10): return Font(color=color, bold=bold, name="Consolas", size=size)
    def border():
        s = Side(style='thin', color='2A2A2A')
        return Border(left=s, right=s, top=s, bottom=s)
    def center(): return Alignment(horizontal='center', vertical='center', wrap_text=False)
    
    # ════════════════════════════════════════
    #  Sheet 1: SELL SIGNALS
    # ════════════════════════════════════════
    ws = wb.active
    ws.title = "SELL SIGNALS"
    ws.sheet_view.showGridLines = False
    
    # Title
    ws.merge_cells('A1:P1')
    t = ws['A1']
    t.value = "⚡ PROFESSIONAL SELL SIGNAL SCREENER  |  10-Indicator Mathematical Analysis"
    t.font = Font(color=C['red_strong'], bold=True, name="Consolas", size=13)
    t.fill = fill(C['bg_dark'])
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32
    
    ws.merge_cells('A2:P2')
    s = ws['A2']
    s.value = f"  Scan Date: {datetime.now().strftime('%d %b %Y  %H:%M')}  |  Indicators: RSI · MACD · Bollinger · Supertrend · EMA · Volume · Stochastic · CISD · Candle · ADX"
    s.font = Font(color=C['grey'], name="Consolas", size=8)
    s.fill = fill(C['bg_panel'])
    s.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 18
    
    # Headers
    headers = [
        "SYMBOL", "COMPANY", "SECTOR", "SIGNAL", "SCORE\n/10",
        "ENTRY\n(₹)", "STOP LOSS\n(₹)", "TARGET 1\n(₹)", "TARGET 2\n(₹)",
        "R:R\nRATIO", "ATR\n(₹)", "RSI",
        "MACD", "SUPER\nTREND", "CISD", "CANDLE"
    ]
    
    col_widths = [12, 22, 14, 17, 8, 10, 12, 10, 10, 8, 8, 7, 16, 16, 16, 20]
    
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font = Font(color=C['red_soft'], bold=True, name="Consolas", size=9)
        cell.fill = fill(C['header_bg'])
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 28
    
    # Data rows
    def write_stock_row(ws, row_idx, r):
        bg = C['row_odd'] if row_idx % 2 == 0 else C['row_even']
        sig = r['signals']
        
        score = r['score']
        score_color = C['red_strong'] if score >= 7 else C['orange'] if score >= 5 else C['yellow'] if score >= 3.5 else C['grey']
        
        row_data = [
            r['symbol'],
            r['company'],
            r['sector'],
            r['signal'],
            score,
            r['entry'],
            r['stop_loss'],
            r['target1'],
            r['target2'],
            r['rr_ratio'],
            r['atr'],
            r['rsi'],
            sig.get('MACD', ''),
            sig.get('Supertrend', ''),
            sig.get('CISD', ''),
            sig.get('Candle', ''),
        ]
        
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=ci, value=val)
            cell.fill = fill(bg)
            cell.border = border()
            cell.alignment = center()
            
            if ci == 1:   cell.font = Font(color=C['yellow'], bold=True, name="Consolas", size=10)
            elif ci == 2: cell.font = font(C['white'], size=9)
            elif ci == 3: cell.font = font(C['grey'], size=9)
            elif ci == 4: cell.font = Font(color=score_color, bold=True, name="Consolas", size=10)
            elif ci == 5: cell.font = Font(color=score_color, bold=True, name="Consolas", size=11)
            elif ci in [6,7,8,9]: cell.font = font(C['white'])
            elif ci == 7: cell.font = font(C['red_soft'])
            elif ci in [8,9]: cell.font = font(C['green'])
            elif ci == 10:
                rr_color = C['green'] if r['rr_ratio'] >= 1.5 else C['yellow']
                cell.font = Font(color=rr_color, bold=True, name="Consolas", size=10)
            elif ci == 12:
                rsi_color = C['red_strong'] if r['rsi'] > 70 else C['orange'] if r['rsi'] > 65 else C['grey']
                cell.font = Font(color=rsi_color, bold=True, name="Consolas", size=10)
            else: cell.font = font(C['grey'], size=9)
        
        ws.row_dimensions[row_idx].height = 22
    
    if sell_stocks:
        for ri, r in enumerate(sell_stocks, 4):
            write_stock_row(ws, ri, r)
    else:
        ws.merge_cells('A4:P4')
        ws['A4'].value = "  ⚪  No stocks crossed the sell threshold today. Market may be trending bullish."
        ws['A4'].font = font(C['grey'])
        ws['A4'].fill = fill(C['bg_panel'])
        ws['A4'].alignment = Alignment(horizontal='center')
    
    # ════════════════════════════════════════
    #  Sheet 2: INDICATOR LEGEND
    # ════════════════════════════════════════
    ws2 = wb.create_sheet("Indicator Guide")
    ws2.sheet_view.showGridLines = False
    
    ws2.merge_cells('A1:D1')
    ws2['A1'].value = "📊 INDICATOR GUIDE & SCORING SYSTEM"
    ws2['A1'].font = Font(color=C['yellow'], bold=True, name="Consolas", size=12)
    ws2['A1'].fill = fill(C['bg_dark'])
    ws2['A1'].alignment = Alignment(horizontal='center')
    ws2.row_dimensions[1].height = 28
    
    guide = [
        ("INDICATOR", "SELL CONDITION", "MAX SCORE", "INTERPRETATION"),
        ("RSI (14)", "Value > 70 + Falling", "1.5 pts", "Overbought zone లో reversal start"),
        ("MACD (12,26,9)", "MACD line crosses below Signal", "1.5 pts", "Momentum bearish గా మారింది"),
        ("Bollinger Bands", "Price rejects upper band (%B > 0.95)", "1.0 pts", "Price mean కి return అవుతుంది"),
        ("Supertrend (10,3)", "Direction flips from Bull to Bear", "1.5 pts", "Trend confirmation — strongest signal"),
        ("EMA Cross (9/21)", "9 EMA crosses below 21 EMA", "1.5 pts", "Short-term momentum కోల్పోయింది"),
        ("Volume", "High volume on Red candle (>1.5x avg)", "1.0 pts", "Institutional selling pressure"),
        ("Stochastic (14,3)", "%K crosses below %D above 80", "1.0 pts", "Overbought exit signal"),
        ("CISD Break", "Displacement candle + Level breakdown", "1.5 pts", "Smart money reversal confirmed"),
        ("Candle Pattern", "Shooting Star / Bearish Engulfing", "1.0 pts", "Price action reversal signal"),
        ("ADX (14)", "ADX > 25 with -DI > +DI", "1.0 pts", "Bearish trend లో strength ఉంది"),
        ("IFC Pattern", "Swing High rejection after funding rally", "2.0 pts", "Institutional reversal — strongest short signal"),
        ("", "", "", ""),
        ("SCORE LEGEND", "", "", ""),
        ("7.0 – 10.0", "STRONG SELL 🔴", "", "High conviction — trade చేయవచ్చు"),
        ("5.0 – 6.9",  "SELL 🟠",        "", "Moderate conviction — caution తో trade"),
        ("3.5 – 4.9",  "WEAK SELL 🟡",   "", "Low conviction — avoid లేదా small size"),
        ("0.0 – 3.4",  "NO SIGNAL ⚪",   "", "Trade వేయకండి"),
        ("", "", "", ""),
        ("TRADE SETUP", "", "", ""),
        ("Entry",       "Today's closing price", "", "Signal day close లో enter అవ్వండి"),
        ("Stop Loss",   "5-day High + 0.5 × ATR", "", "ఇక్కడ కోత పెట్టాలి"),
        ("Target 1",    "Entry − 1.5 × ATR", "", "First profit booking point"),
        ("Target 2",    "Entry − 3.0 × ATR", "", "Full exit point"),
        ("R:R Ratio",   "> 1.5 preferred", "", "Risk కంటే Reward ఎక్కువ ఉండాలి"),
    ]
    
    col_w2 = [20, 35, 12, 40]
    for ci, w in enumerate(col_w2, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w
    
    for ri, row in enumerate(guide, 2):
        for ci, val in enumerate(row, 1):
            cell = ws2.cell(row=ri, column=ci, value=val)
            cell.fill = fill(C['bg_dark'])
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            if ri == 2:
                cell.font = Font(color=C['red_soft'], bold=True, name="Consolas", size=9)
            elif row[0] in ("SCORE LEGEND", "TRADE SETUP"):
                cell.font = Font(color=C['yellow'], bold=True, name="Consolas", size=10)
            elif "STRONG" in str(row[1]):
                cell.font = Font(color=C['red_strong'], name="Consolas", size=9)
            else:
                cell.font = Font(color=C['grey'] if ci == 4 else C['white'], name="Consolas", size=9)
        ws2.row_dimensions[ri].height = 20
    
    # ════════════════════════════════════════
    #  Sheet 3: ALL STOCKS SCAN RESULT
    # ════════════════════════════════════════
    ws3 = wb.create_sheet("All Stocks Summary")
    ws3.sheet_view.showGridLines = False
    
    ws3.merge_cells('A1:F1')
    ws3['A1'].value = "📋 ALL STOCKS — SCAN SUMMARY"
    ws3['A1'].font = Font(color=C['yellow'], bold=True, name="Consolas", size=12)
    ws3['A1'].fill = fill(C['bg_dark'])
    ws3['A1'].alignment = Alignment(horizontal='center')
    ws3.row_dimensions[1].height = 28
    
    h3 = ["SYMBOL", "SIGNAL", "SCORE /10", "ENTRY (₹)", "STOP LOSS (₹)", "R:R RATIO"]
    for ci, h in enumerate(h3, 1):
        cell = ws3.cell(row=2, column=ci, value=h)
        cell.font = Font(color=C['red_soft'], bold=True, name="Consolas", size=9)
        cell.fill = fill(C['header_bg'])
        cell.alignment = center()
    
    all_stocks = sorted(sell_stocks + hold_stocks, key=lambda x: x['score'], reverse=True)
    
    for ri, r in enumerate(all_stocks, 3):
        bg = C['row_odd'] if ri % 2 == 0 else C['row_even']
        score = r['score']
        score_color = C['red_strong'] if score >= 7 else C['orange'] if score >= 5 else C['yellow'] if score >= 3.5 else C['grey']
        
        for ci, val in enumerate([r['symbol'], r['signal'], score, r['entry'], r['stop_loss'], r['rr_ratio']], 1):
            cell = ws3.cell(row=ri, column=ci, value=val)
            cell.fill = fill(bg)
            cell.alignment = center()
            cell.font = Font(color=score_color if ci in [2, 3] else C['white'], name="Consolas", size=9)
    
    for ci, w in enumerate([12, 17, 12, 12, 14, 10], 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w
    
    wb.save(output_path)
    print(f"\n✅ Report saved → {output_path}")


# ══════════════════════════════════════════════════════
#                       MAIN
# ══════════════════════════════════════════════════════

def read_stocks(filepath):
    try:
        df = pd.read_excel(filepath)
        for col in ['Symbol', 'symbol', 'Stock', 'Ticker', 'SYMBOL', 'Name']:
            if col in df.columns:
                symbols = df[col].dropna().astype(str).str.strip().str.upper().tolist()
                return [s for s in symbols if s and s != 'NAN']
        # First column fallback
        symbols = df.iloc[:, 0].dropna().astype(str).str.strip().str.upper().tolist()
        return [s for s in symbols if s and s != 'NAN']
    except Exception as e:
        print(f"❌ Excel read error: {e}")
        return []


def main():
    parser = argparse.ArgumentParser(description='Professional Sell Signal Screener')
    parser.add_argument('--input',  '-i', required=True, help='Input Excel file')
    parser.add_argument('--output', '-o', default='sell_signals.xlsx', help='Output file')
    parser.add_argument('--exchange', '-e', default='NSE', choices=['NSE', 'BSE', 'US'])
    parser.add_argument('--period', '-p', default='3mo', choices=['1mo','3mo','6mo','1y'])
    parser.add_argument('--min-score', '-s', type=float, default=3.5, help='Min score for SELL sheet')
    args = parser.parse_args()
    
    print("\n" + "═"*62)
    print("  ⚡ PROFESSIONAL SELL SIGNAL SCREENER")
    print("     10-Indicator Mathematical Analysis System")
    print("═"*62)
    
    symbols = read_stocks(args.input)
    if not symbols:
        print("❌ No stocks found. Exiting."); return
    
    suffix = {'NSE': '.NS', 'BSE': '.BO', 'US': ''}[args.exchange]
    
    sell_stocks, hold_stocks, errors = [], [], {}
    
    print(f"\n🔍 Scanning {len(symbols)} stocks [{args.exchange}]...\n")
    
    for i, sym in enumerate(symbols, 1):
        ticker_sym = sym + suffix if not sym.endswith(suffix) else sym
        display = sym
        
        print(f"  [{i:3d}/{len(symbols)}] {display:<14}", end=" │ ")
        
        result, err = analyze_stock(ticker_sym, args.period)
        
        if result:
            score = result['score']
            bar = "█" * int(score) + "░" * (10 - int(score))
            print(f"Score: {score:4.1f}/10  [{bar}]  {result['signal']}")
            
            if score >= args.min_score:
                sell_stocks.append(result)
            else:
                hold_stocks.append(result)
        else:
            errors[display] = err
            print(f"⚠️  Error: {err[:50]}")
    
    # Sort by score descending
    sell_stocks.sort(key=lambda x: x['score'], reverse=True)
    
    # Print summary
    print("\n" + "═"*62)
    print(f"  📊 SCAN COMPLETE")
    print(f"     🔴 Sell Signals : {len(sell_stocks)}")
    print(f"     ⚪ No Signal    : {len(hold_stocks)}")
    print(f"     ⚠️  Errors       : {len(errors)}")
    print("═"*62)
    
    if sell_stocks:
        print(f"\n  🔴 TOP SELL CANDIDATES:")
        for r in sell_stocks[:5]:
            print(f"     {r['symbol']:<12} Score:{r['score']:4.1f}  "
                  f"Entry:{r['entry']:<8} SL:{r['stop_loss']:<8} "
                  f"T1:{r['target1']:<8} R:R {r['rr_ratio']}")
    
    create_excel_report(sell_stocks, hold_stocks, args.output)
    print(f"\n  📁 Open '{args.output}' for complete analysis!\n")


if __name__ == "__main__":
    main()
